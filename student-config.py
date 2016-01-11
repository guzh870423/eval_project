import sys
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database_setup import Student, Base, Groups, Semester, Group_Student, Enrollment, Evaluation, EncryptedEvaluation
from ConfigParser import SafeConfigParser
from encrypt import EvalCipher
from sqlalchemy import func, and_
import csv
import openpyxl

parser = SafeConfigParser()
parser.read('config.ini')
username = parser.get('login', 'username')
password = parser.get('login', 'password')
schema = parser.get('login', 'schema')
host = parser.get('login', 'host')
port = parser.get('login', 'port')

engine = create_engine('mysql://' + username + ':' + password + '@' + host +':' + port + '/' + schema) 
# Bind the engine to the metadata of the Base class so that the
# declaratives can be accessed through a DBSession instance
Base.metadata.bind = engine

DBSession = sessionmaker(bind=engine)
# A DBSession() instance establishes all conversations with the database
# and represents a "staging zone" for all the objects loaded into the
# database session object. Any change made against the objects in the
# session won't be persisted into the database until you call
# session.commit(). If you're not happy about the changes, you can
# revert all of them back to the last commit by calling
# session.rollback()
session = DBSession()

wb = openpyxl.load_workbook('student-config.xlsx')
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('students')
rows = sheet.max_row
columns = sheet.max_column
print 'Populating Student configuration data...'

sem_id = sheet['A' + str(2)].value
semester = session.query(Semester).filter_by(id=sem_id).first()
if semester == None:
    print('Invalid Semester ID specified.')
    sys.exit(5)

try:
    enrollment = []
    is_overwrite = False
    for row in range(2, rows + 1):
        semester_id = sheet['A' + str(row)].value
        user_name = sheet['B' + str(row)].value
        first_name = sheet['C' + str(row)].value
        last_name = sheet['D' + str(row)].value
        email = sheet['E' + str(row)].value
        alias_name = sheet['F' + str(row)].value
        
        is_student_exists = session.query(Student).filter_by(user_name=user_name).count()
        is_student_enrolled = session.query(Enrollment).filter_by(student_id=user_name, semester_id=semester_id).count()
        
        student = None
        if is_student_exists > 0:
            student = session.query(Student).filter_by(user_name=user_name).first()
            student.first_name = first_name
            student.last_name = last_name
            student.email = email
            student.alias_name = alias_name
        else:
            student = Student(user_name=user_name, first_name=first_name, last_name=last_name, email=email, alias_name=alias_name)
            session.add(student)
            
        if is_overwrite == False and is_student_enrolled > 0:
            var = raw_input("One or more students already exist in the specified SEMESTER. Would you like to overwrite the ENROLLMENT config? (Y/N): ")
            if var == 'y' or var == 'Y':
                var2 = raw_input("Are you sure you want to overwrite the data? (Y/N): ")
                if var2 == 'y' or var2 == 'Y':    
                    is_overwrite = True
                    session.query(Enrollment).filter_by(semester_id=semester_id).delete()
                elif var2 == 'n' or var2 == 'N':
                    print 'You selected NO.'
                    sys.exit(0)
                else:
                    print 'Invalid choice.'
                    sys.exit(3)    
            elif var == 'n' or var == 'N':
                print 'You selected NO.'
                sys.exit(1)
            else:
                print 'Invalid choice.'
                sys.exit(2)
        
        enrollment.append(Enrollment(student=student, semester=semester))
        
    for enroll in enrollment:
        session.add(enroll)    
    session.commit()
    print "Student configuration tables successfully populated."
except Exception as e:
    print "Error populating student configuration tables.", str(e)
