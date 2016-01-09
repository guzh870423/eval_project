import sys
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database_setup import Student, Base, Groups, Semester, Group_Student, Enrollment, Evaluation, EncryptedEvaluation
from ConfigParser import SafeConfigParser
from encrypt import EvalCipher
from sqlalchemy import func, and_
import csv
import openpyxl

parser = SafeConfigParser()
parser.read('config.ini')
username = parser.get('login', 'username')
password = parser.get('login', 'password')
schema = parser.get('login', 'schema')
host = parser.get('login', 'host')
port = parser.get('login', 'port')

engine = create_engine('mysql://' + username + ':' + password + '@' + host +':' + port + '/' + schema) 
# Bind the engine to the metadata of the Base class so that the
# declaratives can be accessed through a DBSession instance
Base.metadata.bind = engine

DBSession = sessionmaker(bind=engine)
# A DBSession() instance establishes all conversations with the database
# and represents a "staging zone" for all the objects loaded into the
# database session object. Any change made against the objects in the
# session won't be persisted into the database until you call
# session.commit(). If you're not happy about the changes, you can
# revert all of them back to the last commit by calling
# session.rollback()
session = DBSession()

wb = openpyxl.load_workbook('group-config.xlsx')
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('groups')
rows = sheet.max_row
columns = sheet.max_column
print 'Populating configuration data...'

try:
    current_semester_id = sheet['A' + str(2)].value
    current_semester = session.query(Semester).filter_by(id=current_semester_id).first()
    current_week = sheet['B' + str(2)].value
    
    data_count = session.query(Groups).filter_by(semester=current_semester, week=current_week).count()
    if data_count > 0:
        var = raw_input("Matching week's data already present in the database. Would you like to overwrite it? (Y/N): ")
        if var == 'y' or var == 'Y':
            var2 = raw_input("Are you sure you want to overwrite the data? (Y/N): ")
            if var2 == 'y' or var2 == 'Y':    
                session.query(Groups).filter_by(semester=current_semester, week=current_week).delete()
            elif var2 == 'n' or var2 == 'N':
                print 'You selected NO.'
                sys.exit(0)
            else:
                print 'Invalid choice.'
                sys.exit(3)    
        elif var == 'n' or var == 'N':
            print 'You selected NO.'
            sys.exit(1)
        else:
            print 'Invalid choice.'
            sys.exit(2)
except Exception as e:
    print "Unexpected Error occurred.", str(e)    

try:
    for row in range(2, rows + 1):
        semester_id = sheet['A' + str(row)].value
        week = sheet['B' + str(row)].value
        assignment_name = sheet['C' + str(row)].value
        group_name = sheet['D' + str(row)].value
        student_id = sheet['E' + str(row)].value
        is_manager = sheet['F' + str(row)].value
                
        semester = session.query(Semester).filter_by(id=semester_id).first()
        group = session.query(Groups).filter_by(week=week, name=group_name).first()
        if group == None:
            group = Groups(semester=semester, week=week, name=group_name, assignment_name=assignment_name)
            session.add(group)
        student = session.query(Student).filter_by(user_name=student_id).first()
        group_student = Group_Student(groups=group, student=student, is_manager = is_manager)
        session.add(group_student)

    session.commit()
    print "Configuration Data inserted Successfully."
    session.close()
except Exception as e:
    print "Error populating group configuration tables.", str(e)

