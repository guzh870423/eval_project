import sys
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database_setup import Student, Base, Groups, Semester, Group_Student, Enrollment, Evaluation, EncryptedEvaluation
from ConfigParser import SafeConfigParser
from encrypt import EvalCipher
from sqlalchemy import func, and_
import csv
import openpyxl

parser = SafeConfigParser()
parser.read('config.ini')
username = parser.get('login', 'username')
password = parser.get('login', 'password')
schema = parser.get('login', 'schema')
host = parser.get('login', 'host')
port = parser.get('login', 'port')

engine = create_engine('mysql://' + username + ':' + password + '@' + host +':' + port + '/' + schema) 
# Bind the engine to the metadata of the Base class so that the
# declaratives can be accessed through a DBSession instance
Base.metadata.bind = engine

DBSession = sessionmaker(bind=engine)
# A DBSession() instance establishes all conversations with the database
# and represents a "staging zone" for all the objects loaded into the
# database session object. Any change made against the objects in the
# session won't be persisted into the database until you call
# session.commit(). If you're not happy about the changes, you can
# revert all of them back to the last commit by calling
# session.rollback()
session = DBSession()

wb = openpyxl.load_workbook('student-config.xlsx')
wb.get_sheet_names()
sheet = wb.get_sheet_by_name('semester')
rows = sheet.max_row
columns = sheet.max_column
print 'Populating Semester configuration data...'

try:
    semesters = []
    is_overwrite = False
    for row in range(2, rows + 1):
        year = sheet['A' + str(row)].value
        season = sheet['B' + str(row)].value
        course_no = sheet['C' + str(row)].value
        
        is_semester_exists = session.query(Semester).filter_by(year=year, season=season, course_no=course_no).count()
        
        if is_overwrite == False and is_semester_exists > 0:
            var = raw_input("One or more semesters already exist. Would you like to overwrite? (Y/N): ")
            if var == 'y' or var == 'Y':
                var2 = raw_input("Are you sure you want to overwrite the data? (Y/N): ")
                if var2 == 'y' or var2 == 'Y':    
                    is_overwrite = True
                    session.query(Semester).delete()
                elif var2 == 'n' or var2 == 'N':
                    print 'You selected NO.'
                    sys.exit(0)
                else:
                    print 'Invalid choice.'
                    sys.exit(3)    
            elif var == 'n' or var == 'N':
                print 'You selected NO.'
                sys.exit(1)
            else:
                print 'Invalid choice.'
                sys.exit(2)
        
        semesters.append(Semester(year=year, season=season, course_no=course_no))
        
    for sem in semesters:    
        session.add(sem)    
    session.commit()
    session.close()
    print "Semester configuration successfully populated."
except Exception as e:
    print "Error populating Semester configuration.", str(e)
